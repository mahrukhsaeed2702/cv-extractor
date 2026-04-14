"""
CV to Excel Extractor — Streamlit App (Tesseract OCR)
------------------------------------------------------
Usage:
    streamlit run cv_extractor_app.py

Requirements:
    pip install streamlit pdfplumber groq openpyxl pdf2image pillow pytesseract
    System: sudo apt-get install tesseract-ocr poppler-utils
"""

import io
import os
import json
import re

import pdfplumber
import pytesseract
from pdf2image import convert_from_bytes
from PIL import Image
from groq import Groq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import streamlit as st

# ── Configuration ─────────────────────────────────────────────────────────────

GROQ_API_KEY = st.secrets["GROQ_API_KEY"]
GROQ_MODEL   = "llama-3.1-8b-instant"
MAX_TOKENS   = 2048

FIELDS = [
    "Name",
    "Email Address",
    "Phone Number",
    "Address",
    "Highest Degree",
    "University Name",
    "Degree Title",
    "GPA",
    "Specialization",
    "Skills / Tech Stack",
    "Certifications",
    "Experience",
    "Position Applied For",
    "Filename",
]

SYSTEM_PROMPT = """You are a CV/resume parser. Extract structured information from the provided resume text.
Return ONLY a valid JSON object with exactly these keys:
- Name
- Email Address
- Phone Number
- Address (city, country or just city like Lahore)
- Highest Degree (mention if in progress or expected)
- University Name (where highest degree was obtained or is in progress)
- Degree Title (like BS/MS with field e.g. CS, DS, IT, SE)
- GPA
- Specialization (frontend, backend, full stack, AI/ML engineer, type of manager, etc.; null if not mentioned)
- Skills / Tech Stack (comma-separated list)
- Certifications (complete detail, comma-separated)
- Experience (brief summary of all roles with tenure and company names)
- Position Applied For (if clearly stated, otherwise null)

If a field is not found, use an empty string "".
Do NOT include any explanation or markdown — pure JSON only."""

# ── Excel Styles ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
CELL_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")

COL_WIDTHS = {
    "Name": 22, "Email Address": 28, "Phone Number": 18,
    "Address": 30, "Highest Degree": 18, "University Name": 35,
    "Degree Title": 25, "GPA": 8, "Specialization": 22,
    "Skills / Tech Stack": 40, "Certifications": 35,
    "Experience": 45, "Position Applied For": 25, "Filename": 30,
}

# ── Text Extraction ───────────────────────────────────────────────────────────

def extract_text_pdfplumber(pdf_bytes: bytes) -> str:
    """Fast path for text-based PDFs."""
    parts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                parts.append(text.strip())
    return "\n".join(parts).strip()


def extract_text_tesseract(pdf_bytes: bytes) -> str:
    """OCR path for scanned/image-based PDFs (CamScanner etc.)."""
    images = convert_from_bytes(pdf_bytes, dpi=300)
    parts  = []
    for img in images:
        text = pytesseract.image_to_string(img, lang="eng")
        if text.strip():
            parts.append(text.strip())
    return "\n".join(parts).strip()


def extract_text_from_pdf(pdf_bytes: bytes) -> tuple[str, str]:
    """
    Try pdfplumber first (fast).
    Fall back to Tesseract OCR if text is too short (scanned PDF).
    Returns (text, method_used).
    """
    text = extract_text_pdfplumber(pdf_bytes)
    if len(text) >= 50:
        return text, "pdfplumber"

    text = extract_text_tesseract(pdf_bytes)
    return text, "Tesseract OCR"


# ── Groq Extraction ───────────────────────────────────────────────────────────

def extract_fields_with_groq(client: Groq, cv_text: str) -> dict:
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            max_tokens=MAX_TOKENS,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": f"Resume Text:\n\n{cv_text[:6000]}"},
            ],
        )
        raw = response.choices[0].message.content.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        result = json.loads(raw)
        return result if isinstance(result, dict) else {}
    except Exception:
        return {}


# ── Excel Builder ─────────────────────────────────────────────────────────────

def build_excel(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CV Data"

    for col_idx, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row_idx, row_data in enumerate(rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, field in enumerate(FIELDS, start=1):
            value = row_data.get(field, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = CELL_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    for col_idx, field in enumerate(FIELDS, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = COL_WIDTHS.get(field, 20)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────

st.set_page_config(page_title="CV Extractor", page_icon="📄", layout="wide")

with st.sidebar:
    st.image("https://img.icons8.com/fluency/96/resume.png", width=72)
    st.title("CV Extractor")
    st.markdown("Extract structured data from PDF resumes and export to Excel.")
    st.divider()
    st.caption("**Fields extracted:**")
    for f in FIELDS[:-1]:
        st.caption(f"• {f}")
    st.divider()
    st.caption("🔍 Text-based PDFs → pdfplumber (fast)")
    st.caption("🖼️ Scanned PDFs → Tesseract OCR (fallback)")

st.header("📂 Upload PDF Resumes")
st.markdown("Upload one or more CV/resume PDFs — both text-based and scanned (CamScanner) are supported.")

uploaded_files = st.file_uploader(
    "Drop PDFs here or click to browse",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if uploaded_files:
    st.success(f"**{len(uploaded_files)}** file(s) ready: {', '.join(f.name for f in uploaded_files)}")

st.divider()

run_btn = st.button(
    "⚡ Extract & Generate Excel",
    type="primary",
    disabled=not uploaded_files,
    use_container_width=True,
)

if not uploaded_files:
    st.info("Upload at least one PDF resume above to get started.")

# ── Processing ────────────────────────────────────────────────────────────────

if run_btn and uploaded_files:
    client = Groq(api_key=GROQ_API_KEY)
    rows   = []

    progress_bar = st.progress(0, text="Starting…")
    status_box   = st.empty()
    log_expander = st.expander("Processing log", expanded=True)

    total = len(uploaded_files)

    for i, uploaded_file in enumerate(uploaded_files):
        pct  = int((i / total) * 100)
        name = uploaded_file.name
        progress_bar.progress(pct, text=f"Processing {i+1}/{total}: {name}")
        status_box.info(f"🔍 Extracting text from **{name}**…")

        with log_expander:
            log_ph = st.empty()
            log_ph.markdown(f"**→ {name}**")

        try:
            pdf_bytes       = uploaded_file.read()
            cv_text, method = extract_text_from_pdf(pdf_bytes)

            if not cv_text:
                with log_expander:
                    log_ph.warning(f"⚠️ {name} — No text found")
                extracted = {}
            else:
                status_box.info(f"🤖 Calling Groq LLM for **{name}** (via {method})…")
                extracted = extract_fields_with_groq(client, cv_text)
                with log_expander:
                    log_ph.success(f"✅ {name} — Done via {method}")

        except Exception as e:
            with log_expander:
                log_ph.error(f"❌ {name} — Error: {e}")
            extracted = {}

        row = {}
        for field in FIELDS:
            if field == "Filename":
                row[field] = name
            else:
                value = extracted.get(field, "")
                if not value:
                    for k, v in extracted.items():
                        if k.lower() == field.lower():
                            value = v
                            break
                row[field] = str(value).strip() if value else ""
        rows.append(row)

    progress_bar.progress(100, text="Done!")
    status_box.empty()

    st.divider()
    st.subheader(f"✅ Results — {len(rows)} CV(s) processed")
    st.dataframe(rows, use_container_width=True)

    xlsx_bytes = build_excel(rows)
    st.download_button(
        label="⬇️ Download Excel File",
        data=xlsx_bytes,
        file_name="cv_extracted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
